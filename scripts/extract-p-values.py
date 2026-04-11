#!/usr/bin/env python3
"""
P값 데이터 추출기: 방수명가 견적서 템플릿 xlsx에서
공종별 재료/인건/경비 단가를 구조화하여 CSV로 출력.

출력:
  data/p-value-seed.csv          - 전체 항목 (주 데이터)
  data/p-value-summary.csv       - 공종×평단가 구간별 집계
  data/p-value-lump-templates.csv - 소형평수(식) 전용
  data/p-value-extract-log.json  - 디버깅 로그
"""

import csv
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# ── 경로 설정 ──────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = BASE_DIR / "data" / "templates" / "extracted" / "(방수) 방수명가 견적서(최신)"
OUTPUT_DIR = BASE_DIR / "data"

# ── canonicalize (lib/acdb/canonical.ts 포팅) ──────────────
def canonicalize(name: str) -> str:
    """공종명 정규화: 공백 제거, 차수 정리."""
    s = name.strip()
    # 한글 글자 사이 공백 제거 (반복)
    while True:
        next_s = re.sub(r"([가-힣])\s+([가-힣])", r"\1\2", s)
        if next_s == s:
            break
        s = next_s
    # 숫자 + 공백 + 차/회/년/mm 등 붙이기
    s = re.sub(r"(\d)\s+(차|회|년|mm|m|개)", r"\1\2", s)
    # 다중 공백 단일화
    s = re.sub(r"\s+", " ", s).strip()
    # 최종: 모든 공백 제거
    return re.sub(r"\s+", "", s)


# ── 파일명 파싱 ────────────────────────────────────────────
def parse_folder_name(folder: str) -> str:
    """폴더명에서 평수 구간 추출."""
    if "200평 이상" in folder:
        return "200이상"
    if "100평~200평" in folder or "100평~200평 미만" in folder:
        return "100~200"
    if "50평~100평" in folder or "50평~100평 미만" in folder:
        return "50~100"
    if "50평 미만" in folder:
        return "50미만"
    return "기타"


def parse_filename(filename: str, folder_bucket: str) -> dict:
    """파일명에서 공법, 평단가, 총액 등 추출."""
    result = {
        "method": "미정",
        "overall_price_per_m2": 0,
        "total_lump": 0,
        "is_lump_template": False,
    }

    # 소형평수 총액 패턴: "소형평수 330만원" → 3,300,000
    m = re.search(r"소형평수\s*(\d+)만원?", filename)
    if m:
        result["total_lump"] = int(m.group(1)) * 10000
        result["is_lump_template"] = True
        result["method"] = "복합"  # 소형평수는 복합방수 시트 기준
        return result

    # 공법 추출
    if "복합" in filename:
        result["method"] = "복합"
    elif "우레탄" in filename:
        result["method"] = "우레탄"

    # 평단가 추출: "38,000원" 또는 "38000원"
    m = re.search(r"(\d{1,3}),?(\d{3})원", filename)
    if m:
        price_str = m.group(1) + m.group(2)
        result["overall_price_per_m2"] = int(price_str)

    # 20평 이하 소형: "20평 이하 소형평수_50,000원"
    if "20평 이하" in filename or "20평이하" in filename:
        result["method"] = "복합"  # 20평 이하도 복합 시트 기준

    return result


# ── 셀값 안전 읽기 ─────────────────────────────────────────
def safe_num(val) -> float:
    """셀값을 숫자로 변환. None/빈문자/에러 → 0."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace(",", "")
        try:
            return float(val)
        except ValueError:
            return 0
    return 0


def safe_str(val) -> str:
    """셀값을 문자열로 변환."""
    if val is None:
        return ""
    return str(val).strip()


# ── 단위 정규화 ───────────────────────────────────────────
def normalize_unit(unit: str) -> str:
    """단위 통일: ㎡, m² → m2."""
    unit = unit.strip()
    if unit in ("㎡", "m²", "m ²", "㎡"):
        return "m2"
    return unit


# ── 메인 추출 ─────────────────────────────────────────────
def extract_all():
    xlsx_files = sorted(TEMPLATE_DIR.rglob("*.xlsx"))
    if not xlsx_files:
        print(f"ERROR: xlsx 파일 없음 — {TEMPLATE_DIR}")
        sys.exit(1)

    all_records = []
    lump_records = []
    log = {
        "total_files": 0,
        "total_items": 0,
        "canonical_counts": defaultdict(int),
        "anomalies": [],
        "amount_only_items": [],  # 단가=0, 금액만 직접 기입된 항목 (구조적 특성)
        "parse_failures": [],
    }

    for fpath in xlsx_files:
        rel = fpath.relative_to(TEMPLATE_DIR)
        parts = rel.parts
        folder_name = parts[0] if len(parts) > 1 else ""
        filename = fpath.name
        log["total_files"] += 1

        area_bucket = parse_folder_name(folder_name)
        file_info = parse_filename(filename, area_bucket)

        try:
            wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        except Exception as e:
            log["parse_failures"].append({"file": str(rel), "error": str(e)})
            continue

        # Config 시트 → 단가 시점
        price_date = ""
        config_names = [s for s in wb.sheetnames if s.lower() == "config"]
        if config_names:
            cfg_ws = wb[config_names[0]]
            for row in cfg_ws.iter_rows(min_row=1, max_row=1, values_only=True):
                price_date = str(int(safe_num(row[0]))) if row[0] else ""
                break

        # Sheet2 탐색 (이름이 Sheet2인 시트)
        sheet2_names = [s for s in wb.sheetnames if "Sheet2" in s or "sheet2" in s.lower()]
        if not sheet2_names:
            log["parse_failures"].append({"file": str(rel), "error": "Sheet2 없음"})
            wb.close()
            continue

        ws = wb[sheet2_names[0]]
        for row in ws.iter_rows(min_row=7, max_row=17, values_only=True):
            # row: 0=A(None), 1=B(품명), 2=C(규격), 3=D(단위), 4=E(수량),
            #      5=F(재료단가), 6=G(재료금액), 7=H(인건단가), 8=I(인건금액),
            #      9=J(경비단가), 10=K(경비금액), 11=L(합계단가), 12=M(합계금액)
            #      13=N(비고)
            if row is None or len(row) < 12:
                continue

            raw_name = safe_str(row[1])
            if not raw_name:
                continue

            canonical = canonicalize(raw_name)
            spec = safe_str(row[2])
            unit = normalize_unit(safe_str(row[3]))
            qty = safe_num(row[4])
            mat_up = safe_num(row[5])
            mat_amt = safe_num(row[6])
            labor_up = safe_num(row[7])
            labor_amt = safe_num(row[8])
            exp_up = safe_num(row[9])
            exp_amt = safe_num(row[10])
            total_up = safe_num(row[11])
            total_amt = safe_num(row[12])
            remark = safe_str(row[13]) if len(row) > 13 else ""

            # 합계 단가 역산: total_up=0인데 total_amt>0이고 qty>0
            if total_up == 0 and total_amt > 0 and qty > 0:
                total_up = total_amt / qty

            # 이상치 검증: mat + labor + exp vs total (단가 기준)
            computed = mat_up + labor_up + exp_up
            if total_up > 0 and abs(computed - total_up) > 1:
                # 단가=0이지만 금액만 직접 기입 → 구조적 특성 (사다리차/식 항목 등)
                if computed == 0 and total_amt > 0:
                    log["amount_only_items"].append({
                        "file": str(rel),
                        "item": raw_name,
                        "total_amt": total_amt,
                        "reason": "unit_price=0, amount_only",
                    })
                else:
                    log["anomalies"].append({
                        "file": str(rel),
                        "item": raw_name,
                        "computed": computed,
                        "total_up": total_up,
                        "diff": round(computed - total_up, 2),
                    })

            log["canonical_counts"][canonical] += 1
            log["total_items"] += 1

            record = {
                "source_file": filename,
                "area_bucket": area_bucket,
                "method": file_info["method"],
                "overall_price_per_m2": file_info["overall_price_per_m2"],
                "total_lump": file_info["total_lump"],
                "is_lump_template": file_info["is_lump_template"],
                "price_date": price_date,
                "canonical_name": canonical,
                "raw_name": raw_name,
                "spec": spec,
                "unit": unit,
                "qty": qty,
                "mat_unit_price": mat_up,
                "mat_amount": mat_amt,
                "labor_unit_price": labor_up,
                "labor_amount": labor_amt,
                "exp_unit_price": exp_up,
                "exp_amount": exp_amt,
                "total_unit_price": total_up,
                "total_amount": total_amt,
            }

            all_records.append(record)

            # 소형평수(식) 별도 저장
            if file_info["is_lump_template"]:
                lump_records.append(record)

        wb.close()

    return all_records, lump_records, log


# ── CSV 출력 ──────────────────────────────────────────────
SEED_COLS = [
    "source_file", "area_bucket", "method", "overall_price_per_m2",
    "total_lump", "is_lump_template", "price_date",
    "canonical_name", "raw_name", "spec", "unit", "qty",
    "mat_unit_price", "mat_amount", "labor_unit_price", "labor_amount",
    "exp_unit_price", "exp_amount", "total_unit_price", "total_amount",
]

LUMP_COLS = [
    "source_file", "total_lump", "canonical_name", "raw_name",
    "spec", "unit", "qty",
    "mat_amount", "labor_amount", "exp_amount", "total_amount",
]


def write_seed_csv(records):
    path = OUTPUT_DIR / "p-value-seed.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=SEED_COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(records)
    return path


def write_lump_csv(records):
    path = OUTPUT_DIR / "p-value-lump-templates.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=LUMP_COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(records)
    return path


def write_summary_csv(records):
    """공종별 × 평단가 1000원 구간별 집계."""
    path = OUTPUT_DIR / "p-value-summary.csv"

    # 소형평수(식) 제외
    normal = [r for r in records if not r["is_lump_template"]]

    # bucket 그룹핑
    groups = defaultdict(list)
    for r in normal:
        price = r["overall_price_per_m2"]
        bucket_lo = (int(price) // 1000) * 1000
        bucket_hi = bucket_lo + 999
        bucket_key = f"{bucket_lo}-{bucket_hi}"
        groups[(r["canonical_name"], bucket_key)].append(r)

    summary_rows = []
    for (canon, bucket), items in sorted(groups.items()):
        n = len(items)
        avg_mat = sum(r["mat_unit_price"] for r in items) / n
        avg_labor = sum(r["labor_unit_price"] for r in items) / n
        avg_exp = sum(r["exp_unit_price"] for r in items) / n
        avg_total = sum(r["total_unit_price"] for r in items) / n

        mat_r = (avg_mat / avg_total * 100) if avg_total > 0 else 0
        labor_r = (avg_labor / avg_total * 100) if avg_total > 0 else 0
        exp_r = (avg_exp / avg_total * 100) if avg_total > 0 else 0

        sources = sorted(set(r["source_file"] for r in items))

        summary_rows.append({
            "canonical_name": canon,
            "price_bucket": bucket,
            "sample_count": n,
            "avg_mat_unit_price": round(avg_mat, 1),
            "avg_labor_unit_price": round(avg_labor, 1),
            "avg_exp_unit_price": round(avg_exp, 1),
            "mat_ratio": round(mat_r, 1),
            "labor_ratio": round(labor_r, 1),
            "exp_ratio": round(exp_r, 1),
            "source_files": ",".join(sources),
        })

    summary_cols = [
        "canonical_name", "price_bucket", "sample_count",
        "avg_mat_unit_price", "avg_labor_unit_price", "avg_exp_unit_price",
        "mat_ratio", "labor_ratio", "exp_ratio", "source_files",
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=summary_cols)
        w.writeheader()
        w.writerows(summary_rows)
    return path, len(summary_rows)


def write_log(log_data):
    path = OUTPUT_DIR / "p-value-extract-log.json"
    # defaultdict → dict 변환
    out = {
        "total_files": log_data["total_files"],
        "total_items": log_data["total_items"],
        "canonical_counts": dict(log_data["canonical_counts"]),
        "anomaly_count": len(log_data["anomalies"]),
        "anomalies": log_data["anomalies"],
        "amount_only_count": len(log_data["amount_only_items"]),
        "amount_only_items": log_data["amount_only_items"],
        "parse_failure_count": len(log_data["parse_failures"]),
        "parse_failures": log_data["parse_failures"],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    return path


# ── 실행 & 보고 ───────────────────────────────────────────
def main():
    print("=" * 60)
    print("P값 데이터 추출기 v1.0")
    print("=" * 60)

    if not TEMPLATE_DIR.exists():
        print(f"ERROR: 템플릿 디렉토리 없음 — {TEMPLATE_DIR}")
        sys.exit(1)

    all_records, lump_records, log = extract_all()

    # 이상치 비율 체크 (5% 초과 시 경고) — amount_only는 구조적 특성이므로 제외
    anomaly_rate = len(log["anomalies"]) / log["total_items"] * 100 if log["total_items"] > 0 else 0
    if anomaly_rate > 5:
        print(f"WARNING: 진짜 이상치 비율 {anomaly_rate:.1f}% (> 5%). 확인 필요.")

    # 출력 파일 생성
    p1 = write_seed_csv(all_records)
    p2, summary_count = write_summary_csv(all_records)
    p3 = write_lump_csv(lump_records)
    p4 = write_log(log)

    # ── 보고 ───────────────────────────────────────────────
    print(f"\n총 처리 파일: {log['total_files']}")
    print(f"총 추출 항목: {log['total_items']}")
    print(f"파싱 실패:    {len(log['parse_failures'])}")
    print(f"이상치(진짜): {len(log['anomalies'])} ({anomaly_rate:.1f}%)")
    print(f"금액전용항목: {len(log['amount_only_items'])} (구조적 특성, 사다리차/식 등)")

    # area_bucket 분포
    bucket_dist = defaultdict(int)
    for r in all_records:
        bucket_dist[r["area_bucket"]] += 1
    print("\n[area_bucket 분포]")
    for b, c in sorted(bucket_dist.items()):
        print(f"  {b}: {c}건")

    # 고유 canonical 수
    unique_canonicals = set(r["canonical_name"] for r in all_records)
    print(f"\n고유 공종 수: {len(unique_canonicals)}")

    # TOP 10 공종 빈도
    print("\n[TOP 10 공종 빈도]")
    sorted_counts = sorted(log["canonical_counts"].items(), key=lambda x: -x[1])
    for name, cnt in sorted_counts[:10]:
        print(f"  {name}: {cnt}")

    # 출력 파일
    print(f"\n[출력 파일]")
    print(f"  1. {p1}")
    print(f"  2. {p2} ({summary_count} rows)")
    print(f"  3. {p3} ({len(lump_records)} rows)")
    print(f"  4. {p4}")

    # 소형평수(식) 요약
    if lump_records:
        print(f"\n[소형평수(식) 요약 — {len(lump_records)} 항목]")
        lump_files = sorted(set(r["source_file"] for r in lump_records))
        for lf in lump_files:
            items = [r for r in lump_records if r["source_file"] == lf]
            total = sum(r["total_amount"] for r in items)
            print(f"  {lf}: {len(items)}개 공종, 소계합={total:,.0f}")


if __name__ == "__main__":
    main()
